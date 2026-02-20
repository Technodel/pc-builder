import streamlit as st
import pandas as pd
import openpyxl
import re
import os

# --- 1. PAGE CONFIG & LOGO ---
st.set_page_config(page_title="Technodel Pro Builder", layout="wide", page_icon="💻")

# Logo: Left-aligned
st.markdown(
    """
    <div style="text-align: left; padding-bottom: 5px;">
        <img src="https://technodel.net/wp-content/uploads/2024/08/technodel-site-logo-01.webp" width="280">
    </div>
    """,
    unsafe_allow_html=True
)

st.markdown("""
    <style>
    .stSelectbox { margin-bottom: -15px; }
    .stButton button { width: 100%; border-radius: 5px; height: 2.2em; margin-top: 10px;}
    .live-summary {
        background-color: #f0f8ff;
        padding: 20px;
        border-radius: 10px;
        border: 1px solid #00a8e8;
        margin-bottom: 20px;
    }
    .screenshot-box { 
        background-color: #ffffff; 
        padding: 30px; 
        border: 2px solid #333; 
        border-radius: 12px; 
        font-family: 'Segoe UI', sans-serif; 
        color: #000000;
        margin-top: 20px;
    }
    .contact-footer { 
        background-color: #f8f9fa; 
        padding: 20px; 
        border-radius: 10px; 
        border-left: 5px solid #00a8e8;
        margin-top: 30px;
    }
    .social-icon { width: 18px; height: 18px; vertical-align: middle; margin-right: 8px; }
    </style>
    """, unsafe_allow_html=True)

# --- 2. HEADER: TITLE & FILE SELECTOR ---
def get_all_excel_files():
    base_path = os.path.dirname(os.path.abspath(__file__))
    files = [f for f in os.listdir(base_path) if f.endswith(('.xlsx', '.xlsm'))]
    if not files: return []
    full_paths = [os.path.join(base_path, f) for f in files]
    full_paths.sort(key=os.path.getmtime, reverse=True)
    return [os.path.basename(p) for p in full_paths]

head_col1, head_col2 = st.columns([2, 1])

with head_col1:
    st.title("💻 Technodel PC Builder Pro")

with head_col2:
    excel_options = get_all_excel_files()
    if excel_options:
        selected_filename = st.selectbox("📂 Database Source", excel_options)
        FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), selected_filename)
    else:
        st.error("❌ No Excel files found!")
        st.stop()

# --- 3. DATA LOADER ---
@st.cache_data
def load_data(file_path, keyword):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_map = {s.upper(): s for s in wb.sheetnames}
        if "HARDWARE" not in sheet_map: return pd.DataFrame()
        ws = wb[sheet_map["HARDWARE"]]
        data = []
        target_table = next((ws.tables[t] for t in ws.tables if keyword.lower() in t.lower()), None)
        rows = ws[target_table.ref] if target_table else list(ws.rows)
        found_section = True if target_table else False
        current_ram_tech = None 

        for row in rows:
            val_a = str(row[0].value).strip() if row[0].value else ""
            if not found_section:
                if keyword.lower() in val_a.lower(): found_section = True
                continue
            if not val_a or val_a == "None":
                if target_table: continue
                else: break
            if "DDR4" in val_a.upper(): current_ram_tech = "DDR4"
            if "DDR5" in val_a.upper(): current_ram_tech = "DDR5"
            raw_price = row[1].value if len(row) > 1 else 0
            try:
                clean_price = int(round(float(raw_price), 0))
                if clean_price <= 0: continue
            except: continue
            data.append({
                "ITEM": val_a, "PRICE": clean_price,
                "RAM_TECH": "DDR5" if "DDR5" in val_a.upper() else ("DDR4" if "DDR4" in val_a.upper() else current_ram_tech)
            })
        return pd.DataFrame(data)
    except: return pd.DataFrame()

sections = ["cpu", "mb", "ram", "gpu", "case", "psu", "coo", "storage"]
dfs = {s: load_data(FILE_PATH, s) for s in sections}

# --- 4. COMPATIBILITY ---
def get_cpu_gen(cpu_name):
    match = re.search(r'(\d{4,5})', cpu_name)
    return (match.group(1)[:2] if len(match.group(1)) == 5 else match.group(1)[0]) if match else None

def is_compat(cpu_sel, mb_name):
    if "Select" in cpu_sel: return True
    gen = get_cpu_gen(cpu_sel)
    match = re.search(r'\((.*?)\)', mb_name)
    return str(gen) in re.findall(r'\d+', match.group(1)) if (gen and match) else True

# --- 5. COMPONENT SELECTION ---
col1, col2 = st.columns([1, 1])
with col1:
    cpu_choice = st.selectbox("Processor", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['cpu'].iterrows()], key="c")
    mb_list = dfs['mb']
    if "Select" not in cpu_choice:
        filtered = mb_list[mb_list['ITEM'].apply(lambda x: is_compat(cpu_choice, x))]
        if not filtered.empty: mb_list = filtered
    mb_choice = st.selectbox("Motherboard", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in mb_list.iterrows()], key="m")
    st.divider()
    if "Select" not in mb_choice:
        mb_type = "DDR5" if "DDR5" in mb_choice.upper() else "DDR4"
        ram_list = dfs['ram'][dfs['ram']['RAM_TECH'] == mb_type]
        if 'ram_count' not in st.session_state: st.session_state.ram_count = 1
        for i in range(st.session_state.ram_count):
            st.selectbox(f"RAM {i+1}", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in ram_list.iterrows()], key=f"ram_{i}")
        if st.button("➕ Add RAM"): st.session_state.ram_count += 1; st.rerun()

with col2:
    gpu_choice = st.selectbox("GPU", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['gpu'].iterrows()], key="g")
    psu_choice = st.selectbox("PSU", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['psu'].iterrows()], key="p")
    case_choice = st.selectbox("Case", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['case'].iterrows()], key="ca")
    cool_choice = st.selectbox("Cooling", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['coo'].iterrows()], key="co")
    st.divider()
    if 'storage_count' not in st.session_state: st.session_state.storage_count = 1
    for i in range(st.session_state.storage_count):
        st.selectbox(f"Storage {i+1}", ["Select"] + [f"{r['ITEM']} - ${r['PRICE']}" for _,r in dfs['storage'].iterrows()], key=f"store_{i}")
    if st.button("➕ Add Storage"): st.session_state.storage_count += 1; st.rerun()

# --- 6. LIVE TOTALS & PREVIEW ---
st.divider()
total = 0
items_for_preview = []
items_for_text = []

keys = [('CPU', 'c'), ('Motherboard', 'm'), ('GPU', 'g'), ('PSU', 'p'), ('Case', 'ca'), ('Cooler', 'co')]
for label, k in keys:
    val = st.session_state.get(k)
    if val and "Select" not in val:
        total += int(val.split(" - $")[1].replace(",",""))
        items_for_preview.append(f"<b>{label}:</b> {val}")
        items_for_text.append(f"{label}: {val}")

for i in range(st.session_state.get('ram_count', 1)):
    val = st.session_state.get(f"ram_{i}")
    if val and "Select" not in val:
        total += int(val.split(" - $")[1].replace(",",""))
        items_for_preview.append(f"<b>RAM {i+1}:</b> {val}")
        items_for_text.append(f"RAM {i+1}: {val}")

for i in range(st.session_state.get('storage_count', 1)):
    val = st.session_state.get(f"store_{i}")
    if val and "Select" not in val:
        total += int(val.split(" - $")[1].replace(",",""))
        items_for_preview.append(f"<b>Storage {i+1}:</b> {val}")
        items_for_text.append(f"Storage {i+1}: {val}")

# Live Builder Summary
if items_for_text:
    st.subheader("🛒 Real-Time Build Summary")
    list_items_md = "\n".join([f"* {i}" for i in items_for_text])
    st.markdown(f'<div class="live-summary">{list_items_md}<h2 style="color:#00a8e8; margin-top:15px;">TOTAL: ${total:,}</h2></div>', unsafe_allow_html=True)

# Download Button
summary_txt = "\n".join(items_for_text) + f"\n\nTOTAL: ${total:,}"
st.download_button("📥 Download Summary (.txt)", summary_txt, "technodel_build.txt")

# Order Info Footer
st.markdown(f"""
<div class="contact-footer">
    <h3 style="margin-top:0;">📋 Order Information</h3>
    <p>✅ <b>Warranty:</b> 1 Year warranty on all parts | 🚚 <b>Availability:</b> Ready for pickup within 24 hours</p>
    <p>📞 03659872 | 70449900 | 71234002 | 07345689 | 📧 Adarwich@engineer.com</p>
    <hr>
    <img src="https://cdn-icons-png.flaticon.com/512/733/733547.png" class="social-icon"> fb.com/technodel | 
    <img src="https://cdn-icons-png.flaticon.com/512/2111/2111463.png" class="social-icon"> @technodel_computers
</div>
""", unsafe_allow_html=True)

# Professional Screenshot Box
if st.checkbox("📸 Show Professional Screenshot Box"):
    if items_for_preview:
        build_list_html = "<br>".join(items_for_preview)
        full_preview = f"""
        <div class="screenshot-box">
            <h2 style="margin-top:0; color:#00a8e8;">TECHNODEL BUILD SUMMARY</h2>
            <p style="font-size:1.1em; font-weight:bold;">✅ 1 Year Warranty | 🚚 24h Pickup</p>
            <hr>
            <div style="padding: 10px 0; font-size: 1.1em; line-height: 1.8;">{build_list_html}</div>
            <h2 style="color:#d32f2f; border-top: 2px solid #eee; padding-top: 15px;">TOTAL: ${total:,}</h2>
        </div>
        """
        st.markdown(full_preview, unsafe_allow_html=True)
    else:
        st.warning("⚠️ Select parts first to see the preview.")

if st.button("Reset Everything"):
    for key in list(st.session_state.keys()): del st.session_state[key]
    st.rerun()
